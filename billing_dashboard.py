#!/usr/bin/env python3
"""GEP Billing Dashboard — read-only monitoring for automated billing runs.

Displays run results produced by the billing watcher / billing engine.
No uploads, no manual triggers. Just open it and see what happened.

    streamlit run billing_dashboard.py
"""

from __future__ import annotations

import io
import json
import logging
import subprocess
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from billing_engine import DEFAULT_INPUTS_DIR, DEFAULT_OUTPUTS_DIR

OUTPUTS_DIR = DEFAULT_OUTPUTS_DIR
BILLING_LOG_DIR = OUTPUTS_DIR / "gep_billing_log"
PARTNER_DETAILS_DIR = OUTPUTS_DIR / "gep_partner_details"
NETSUITE_CSV_PATH = OUTPUTS_DIR / "gep_netsuite_invoice_import.csv"

ACCRUAL_OUTPUT_DIR = OUTPUTS_DIR / "gep_accrual"

WATCHER_STATE_DIR = Path("outputs") / "watcher_state"
WATCHER_LEDGER = WATCHER_STATE_DIR / "processed_files.json"
WATCHER_LOG = Path("logs") / "watcher_stderr.log"

logger = logging.getLogger("billing_dashboard")


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------
def _find_manifests() -> List[Dict[str, Any]]:
    """Load all run manifest JSONs from the billing log directory, newest first."""
    if not BILLING_LOG_DIR.exists():
        return []
    manifests = []
    for path in sorted(BILLING_LOG_DIR.glob("*_run_manifest_*.json"), reverse=True):
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
                data["_manifest_path"] = str(path)
                manifests.append(data)
        except Exception:
            continue
    return manifests


def _find_master_reports() -> List[Path]:
    """Return all Master Billing Reports sorted newest first."""
    if not BILLING_LOG_DIR.exists():
        return []
    return sorted(BILLING_LOG_DIR.glob("*_Master_Billing_Report.xlsx"), reverse=True)


def _read_audit_controls(report_path: Path) -> List[Dict[str, Any]]:
    if not report_path.exists():
        return []
    try:
        wb = load_workbook(report_path, data_only=True)
    except Exception:
        return []
    if "Audit & Controls" not in wb.sheetnames:
        return []
    ws = wb["Audit & Controls"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        if values is None:
            continue
        row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        if any(v is not None and str(v).strip() for v in row.values()):
            out.append(row)
    return out


def _read_executive_summary(report_path: Path) -> Optional[pd.DataFrame]:
    if not report_path.exists():
        return None
    try:
        df = pd.read_excel(report_path, sheet_name="Executive Summary", index_col=0)
        return df
    except Exception:
        return None


def _audit_status(audit_rows: List[Dict[str, Any]]) -> str:
    statuses = [str(row.get("Status", "")).upper() for row in audit_rows]
    if any(s == "FAIL" for s in statuses):
        return "FAIL"
    if any(s == "REVIEW REQUIRED" for s in statuses):
        return "REVIEW REQUIRED"
    if any(s == "PASS" for s in statuses):
        return "PASS"
    return "UNKNOWN"


def _billing_period_from_report(report_path: Path) -> str:
    stem = report_path.stem
    if "_" in stem:
        return stem.split("_")[0]
    return "Unknown"


def _load_watcher_ledger() -> Dict[str, Any]:
    if not WATCHER_LEDGER.exists():
        return {}
    try:
        with WATCHER_LEDGER.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _watcher_is_running() -> bool:
    try:
        result = subprocess.run(
            ["launchctl", "list"],
            capture_output=True, text=True, timeout=5,
        )
        return "com.gep.billing-watcher" in result.stdout
    except Exception:
        return False


def _watcher_recent_log(lines: int = 25) -> str:
    if not WATCHER_LOG.exists():
        return "No log file found."
    try:
        all_lines = WATCHER_LOG.read_text(encoding="utf-8", errors="replace").splitlines()
        return "\n".join(all_lines[-lines:]) if all_lines else "Log file is empty."
    except Exception as exc:
        return f"Could not read log: {exc}"


def _zip_partner_details() -> Optional[bytes]:
    if not PARTNER_DETAILS_DIR.exists():
        return None
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in PARTNER_DETAILS_DIR.rglob("*"):
            if file_path.is_file() and not file_path.name.startswith("~$"):
                archive.write(file_path, arcname=file_path.relative_to(PARTNER_DETAILS_DIR))
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
def _render_status_banner(audit_rows: List[Dict[str, Any]], billing_period: str) -> None:
    status = _audit_status(audit_rows)
    if status == "PASS":
        st.success(f"**{billing_period}** — All audit controls passed", icon="\u2705")
    elif status == "FAIL":
        st.error(f"**{billing_period}** — Audit control failure detected", icon="\U0001f6a8")
    elif status == "REVIEW REQUIRED":
        st.warning(f"**{billing_period}** — Review required on one or more controls", icon="\u26a0\ufe0f")
    else:
        st.info(f"**{billing_period}** — Audit status unknown")


def _render_summary_metrics(summary_df: Optional[pd.DataFrame]) -> None:
    if summary_df is None or summary_df.empty:
        return

    total_billed = 0.0
    total_usage = 0.0
    total_next_day = 0.0
    total_minimum = 0.0
    total_end_users = 0
    total_billable_users = 0
    partner_count = len(summary_df)

    for col in summary_df.columns:
        col_lower = col.lower().strip() if isinstance(col, str) else ""
        series = pd.to_numeric(summary_df[col], errors="coerce").fillna(0)
        if "total billed" in col_lower:
            total_billed = series.sum()
        elif "usage revenue" in col_lower:
            total_usage = series.sum()
        elif "next day" in col_lower:
            total_next_day = series.sum()
        elif "minimum revenue" in col_lower:
            total_minimum = series.sum()
        elif "active end user" in col_lower:
            total_end_users = int(series.sum())
        elif "billable" in col_lower and "user" in col_lower:
            total_billable_users = int(series.sum())

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Billed", f"${total_billed:,.2f}")
    col2.metric("Partners", f"{partner_count}")
    col3.metric("Active End Users", f"{total_end_users:,}")
    col4.metric("Billable Users", f"{total_billable_users:,}")

    col5, col6, col7 = st.columns(3)
    col5.metric("Usage Revenue", f"${total_usage:,.2f}")
    col6.metric("Next-Day Fee Revenue", f"${total_next_day:,.2f}")
    col7.metric("Minimum Revenue", f"${total_minimum:,.2f}")


def _render_downloads(report_path: Optional[Path], key_prefix: str) -> None:
    cols = st.columns(4)

    with cols[0]:
        if report_path and report_path.exists():
            st.download_button(
                "Master Billing Report",
                data=report_path.read_bytes(),
                file_name=report_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key_prefix}_master",
                use_container_width=True,
            )

    with cols[1]:
        if NETSUITE_CSV_PATH.exists():
            st.download_button(
                "NetSuite Import CSV",
                data=NETSUITE_CSV_PATH.read_bytes(),
                file_name=NETSUITE_CSV_PATH.name,
                mime="text/csv",
                key=f"{key_prefix}_netsuite",
                use_container_width=True,
            )

    with cols[2]:
        partner_zip = _zip_partner_details()
        if partner_zip:
            period = _billing_period_from_report(report_path) if report_path else "unknown"
            st.download_button(
                "Partner Details (ZIP)",
                data=partner_zip,
                file_name=f"{period}_partner_details.zip",
                mime="application/zip",
                key=f"{key_prefix}_partner",
                use_container_width=True,
            )

    with cols[3]:
        manifests = _find_manifests()
        if manifests:
            latest = manifests[0]
            manifest_path = Path(latest["_manifest_path"])
            if manifest_path.exists():
                st.download_button(
                    "Run Manifest",
                    data=manifest_path.read_bytes(),
                    file_name=manifest_path.name,
                    mime="application/json",
                    key=f"{key_prefix}_manifest",
                    use_container_width=True,
                )


def page_overview() -> None:
    reports = _find_master_reports()

    if not reports:
        st.info("No billing runs found yet. The watcher will create outputs here once a usage file is detected.")
        return

    latest_report = reports[0]
    billing_period = _billing_period_from_report(latest_report)
    audit_rows = _read_audit_controls(latest_report)
    summary_df = _read_executive_summary(latest_report)

    _render_status_banner(audit_rows, billing_period)

    st.markdown("### Key Metrics")
    _render_summary_metrics(summary_df)

    st.markdown("---")
    st.markdown("### Audit & Controls")
    if audit_rows:
        audit_df = pd.DataFrame(audit_rows)
        def _style_status(val):
            if str(val).upper() == "PASS":
                return "background-color: #d4edda; color: #155724"
            if str(val).upper() == "FAIL":
                return "background-color: #f8d7da; color: #721c24"
            if "REVIEW" in str(val).upper():
                return "background-color: #fff3cd; color: #856404"
            return ""
        styled = audit_df.style.map(_style_status, subset=["Status"])
        st.dataframe(styled, use_container_width=True, hide_index=True)
    else:
        st.warning("No audit data found in the latest report.")

    st.markdown("---")
    st.markdown("### Executive Summary by Partner")
    if summary_df is not None and not summary_df.empty:
        st.dataframe(
            summary_df.style.format({
                col: "${:,.2f}" for col in summary_df.columns
                if any(kw in col.lower() for kw in ["revenue", "billed", "minimum"])
            }),
            use_container_width=True,
        )
    else:
        st.info("Executive summary not available.")

    st.markdown("---")
    st.markdown("### Downloads")
    _render_downloads(latest_report, key_prefix="overview")


def page_run_history() -> None:
    manifests = _find_manifests()

    if not manifests:
        st.info("No run history found.")
        return

    table_rows = []
    for m in manifests:
        table_rows.append({
            "Billing Period": m.get("billing_period", ""),
            "Run ID": m.get("run_id", ""),
            "Timestamp (UTC)": m.get("run_timestamp_utc", ""),
            "Status": m.get("status", ""),
            "Usage File": m.get("usage_file", {}).get("name", ""),
            "Git Commit": m.get("engine_git_commit", ""),
        })
    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    reports = _find_master_reports()
    if len(reports) > 1:
        st.markdown("---")
        selected_name = st.selectbox(
            "View details for a specific report",
            options=[r.name for r in reports],
        )
        selected = next((r for r in reports if r.name == selected_name), None)
        if selected:
            period = _billing_period_from_report(selected)
            audit = _read_audit_controls(selected)
            summary = _read_executive_summary(selected)

            _render_status_banner(audit, period)

            if audit:
                audit_df = pd.DataFrame(audit)
                st.dataframe(audit_df, use_container_width=True, hide_index=True)

            if summary is not None and not summary.empty:
                st.markdown("**Executive Summary**")
                st.dataframe(summary, use_container_width=True)

            st.download_button(
                f"Download {selected.name}",
                data=selected.read_bytes(),
                file_name=selected.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"history_dl_{selected.name}",
            )


def page_watcher_status() -> None:
    running = _watcher_is_running()

    if running:
        st.success("Watcher is running", icon="\u2705")
    else:
        st.error("Watcher is not running", icon="\u274c")
        st.caption("Start it with: `launchctl load ~/Library/LaunchAgents/com.gep.billing-watcher.plist`")

    st.markdown("---")
    st.markdown("### Processed Files")
    ledger = _load_watcher_ledger()
    processed = ledger.get("processed", {})

    if not processed:
        st.info("No files have been processed yet.")
    else:
        rows = []
        for filename, info in processed.items():
            rows.append({
                "File": filename,
                "Status": info.get("status", info.get("note", "processed")),
                "Billing Period": info.get("billing_period", ""),
                "Run ID": info.get("run_id", ""),
                "Processed At": info.get("processed_at_utc", info.get("detected_at_utc", "")),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("### Recent Watcher Log")
    log_text = _watcher_recent_log(30)
    st.code(log_text, language="log")

    st.markdown("---")
    st.markdown("### Watcher Configuration")
    st.markdown(f"- **Watching:** `{DEFAULT_INPUTS_DIR}`")
    st.markdown(f"- **Outputs:** `{DEFAULT_OUTPUTS_DIR}`")
    st.markdown("- **Poll interval:** 30 minutes")
    st.markdown(f"- **Ledger:** `{WATCHER_LEDGER}`")


def page_accruals() -> None:
    if not ACCRUAL_OUTPUT_DIR.exists():
        st.info("No accrual outputs found yet. Run the accrual engine to generate results here.")
        st.code(
            'python3 accrual_engine.py accrual \\\n'
            '  --accrual-month 2026-02 \\\n'
            '  --rules-path "/path/to/gep_billing_rules.xlsx"',
            language="bash",
        )
        return

    je_files = sorted(ACCRUAL_OUTPUT_DIR.glob("gep_accrual_JE_*.csv"), reverse=True)
    totals_files = sorted(ACCRUAL_OUTPUT_DIR.glob("gep_accrual_totals_*.csv"), reverse=True)
    variance_files = sorted(ACCRUAL_OUTPUT_DIR.glob("variance_*.csv"), reverse=True)
    detail_files = sorted(ACCRUAL_OUTPUT_DIR.glob("gep_accrual_billing_detail_*.xlsx"), reverse=True)

    if not je_files and not totals_files and not variance_files:
        st.info("Accrual output directory exists but no results found yet.")
        return

    # --- Accrual totals (most useful at a glance) ---
    if totals_files:
        st.markdown("### Accrual Totals")
        latest_totals = totals_files[0]
        period_str = latest_totals.stem.replace("gep_accrual_totals_", "")
        st.caption(f"Period: {period_str[:4]}.{period_str[4:]}")

        totals_df = pd.read_csv(latest_totals)
        grand_total = totals_df["Total"].sum() if "Total" in totals_df.columns else 0

        col1, col2 = st.columns(2)
        col1.metric("Estimated Total Accrual", f"${grand_total:,.2f}")
        col2.metric("Customers", f"{len(totals_df)}")

        format_cols = {c: "${:,.2f}" for c in totals_df.columns if c != "Customer"}
        st.dataframe(
            totals_df.style.format(format_cols),
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            "Download Accrual Totals CSV",
            data=latest_totals.read_bytes(),
            file_name=latest_totals.name,
            mime="text/csv",
            key="dl_accrual_totals",
        )

    # --- Journal Entry ---
    if je_files:
        st.markdown("---")
        st.markdown("### Journal Entry (NetSuite Import)")
        latest_je = je_files[0]
        je_df = pd.read_csv(latest_je)
        st.dataframe(je_df, use_container_width=True, hide_index=True)

        st.download_button(
            "Download JE CSV",
            data=latest_je.read_bytes(),
            file_name=latest_je.name,
            mime="text/csv",
            key="dl_accrual_je",
        )

    # --- Variance report ---
    if variance_files:
        st.markdown("---")
        st.markdown("### Variance Report (Accrual vs Actual)")

        latest_var = variance_files[0]
        var_df = pd.read_csv(latest_var)

        total_row = var_df[var_df["Customer"] == "TOTAL"]
        if not total_row.empty:
            total_var_pct = total_row.iloc[0].get("Variance_Pct", 0)
            total_var_amt = total_row.iloc[0].get("Variance", 0)
            if abs(float(total_var_pct)) > 5:
                st.warning(f"Total variance: ${float(total_var_amt):,.2f} ({float(total_var_pct):.1f}%) — exceeds 5% materiality threshold")
            else:
                st.success(f"Total variance: ${float(total_var_amt):,.2f} ({float(total_var_pct):.1f}%) — within materiality threshold")

        def _style_flag(val):
            if str(val).strip() == "\u26a0":
                return "background-color: #fff3cd; color: #856404"
            if str(val).strip() == "\u2713":
                return "background-color: #d4edda; color: #155724"
            return ""

        format_map = {c: "${:,.2f}" for c in ["Estimated", "Actual", "Variance"] if c in var_df.columns}
        if "Variance_Pct" in var_df.columns:
            format_map["Variance_Pct"] = "{:.1f}%"

        styled = var_df.style.format(format_map)
        if "Flag" in var_df.columns:
            styled = styled.map(_style_flag, subset=["Flag"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        st.download_button(
            "Download Variance Report",
            data=latest_var.read_bytes(),
            file_name=latest_var.name,
            mime="text/csv",
            key="dl_variance",
        )

    # --- Billing detail ---
    if detail_files:
        st.markdown("---")
        st.markdown("### Billing Detail (Fee Calculation Review)")
        latest_detail = detail_files[0]
        st.download_button(
            f"Download {latest_detail.name}",
            data=latest_detail.read_bytes(),
            file_name=latest_detail.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_accrual_detail",
        )

    # --- Historical files ---
    all_files = sorted(
        list(je_files) + list(totals_files) + list(variance_files) + list(detail_files),
        reverse=True,
    )
    if len(all_files) > 1:
        st.markdown("---")
        with st.expander("All accrual files"):
            for f in all_files:
                st.markdown(f"- `{f.name}`")


def main() -> None:
    st.set_page_config(
        page_title="GEP Billing Dashboard",
        page_icon="\U0001f4ca",
        layout="wide",
    )

    st.title("GEP Billing Dashboard")
    st.caption("Automated billing monitoring — no uploads, no buttons. Just results.")

    tab_overview, tab_accruals, tab_history, tab_watcher = st.tabs([
        "Latest Run",
        "Accruals",
        "Run History",
        "Watcher Status",
    ])

    with tab_overview:
        page_overview()

    with tab_accruals:
        page_accruals()

    with tab_history:
        page_run_history()

    with tab_watcher:
        page_watcher_status()


if __name__ == "__main__":
    main()
