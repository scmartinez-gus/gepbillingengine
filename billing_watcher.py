#!/usr/bin/env python3
"""Polling file watcher for automated GEP billing runs.

Monitors the inputs directory for new gepusage*.csv files. When a new file
appears and is confirmed stable (Google Drive sync finished), the billing
engine runs automatically. Results are logged and optionally sent to Slack.

Usage:
    python3 billing_watcher.py                      # defaults (30-min poll)
    python3 billing_watcher.py --poll-interval 600   # every 10 minutes
    python3 billing_watcher.py --dry-run              # detect files but don't run billing
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import signal
import sys
import time
import traceback
import urllib.error
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from billing_engine import (
    BillingEngineError,
    DEFAULT_CONFIG_FILE,
    DEFAULT_INPUTS_DIR,
    DEFAULT_OUTPUTS_DIR,
    DEFAULT_USAGE_PREFIX,
    run_billing_engine,
)
from accrual_engine import (
    DEFAULT_ACCRUAL_OUTPUT_DIR,
    DEFAULT_V3_USAGE_DIR,
    run_accrual,
)

DEFAULT_POLL_INTERVAL = 1800  # 30 minutes
DEFAULT_ACCRUAL_DAY = 25  # day of month to auto-run accrual

WATCHER_STATE_DIR = Path("outputs") / "watcher_state"
PROCESSED_LEDGER = WATCHER_STATE_DIR / "processed_files.json"

logger = logging.getLogger("billing_watcher")


# ---------------------------------------------------------------------------
# Graceful shutdown
# ---------------------------------------------------------------------------
_shutdown_requested = False


def _handle_signal(signum: int, frame: Any) -> None:
    global _shutdown_requested
    _shutdown_requested = True
    logger.info("Shutdown signal received (signal %d). Will exit after current cycle.", signum)


# ---------------------------------------------------------------------------
# Ledger: tracks which files have already been processed
# ---------------------------------------------------------------------------
def _ensure_state_dir() -> None:
    WATCHER_STATE_DIR.mkdir(parents=True, exist_ok=True)


def _load_ledger() -> Dict[str, Any]:
    if not PROCESSED_LEDGER.exists():
        return {"processed": {}}
    try:
        with PROCESSED_LEDGER.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if "processed" not in data:
            data["processed"] = {}
        return data
    except (json.JSONDecodeError, OSError):
        logger.warning("Ledger file corrupt or unreadable — starting fresh.")
        return {"processed": {}}


def _save_ledger(ledger: Dict[str, Any]) -> None:
    _ensure_state_dir()
    tmp = PROCESSED_LEDGER.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(ledger, f, indent=2, sort_keys=True)
    tmp.replace(PROCESSED_LEDGER)


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------
def _find_usage_candidates(inputs_dir: Path, prefix: str) -> List[Path]:
    """Return gepusage*.csv files sorted newest-first by modification time."""
    if not inputs_dir.exists():
        return []
    candidates = [
        p
        for p in inputs_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".csv"
        and p.name.lower().startswith(prefix.lower())
    ]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates


# ---------------------------------------------------------------------------
# File stability check (confirms Google Drive sync is done)
# ---------------------------------------------------------------------------
_pending_files: Dict[str, int] = {}  # path -> file size on first sighting


def _is_file_stable(path: Path) -> bool:
    """Return True when the file size hasn't changed since the previous poll.

    On first sighting the file is recorded and we return False (wait one more
    cycle). On the next poll, if the size matches we consider it fully synced.
    """
    key = str(path)
    try:
        current_size = path.stat().st_size
    except OSError:
        _pending_files.pop(key, None)
        return False

    if current_size == 0:
        return False

    previous_size = _pending_files.get(key)
    if previous_size is None:
        _pending_files[key] = current_size
        logger.info("New file detected: %s (%d bytes). Waiting one cycle to confirm stability.", path.name, current_size)
        return False

    if current_size != previous_size:
        _pending_files[key] = current_size
        logger.info("File %s still syncing (%d → %d bytes). Waiting another cycle.", path.name, previous_size, current_size)
        return False

    _pending_files.pop(key, None)
    logger.info("File %s is stable (%d bytes). Ready to process.", path.name, current_size)
    return True


# ---------------------------------------------------------------------------
# Audit & Controls reader (reused from portal logic)
# ---------------------------------------------------------------------------
def _read_audit_controls(report_path: Path) -> List[Dict[str, Any]]:
    if not report_path.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(report_path, data_only=True)
    except Exception:
        return []
    if "Audit & Controls" not in wb.sheetnames:
        return []
    ws = wb["Audit & Controls"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        if values is None:
            continue
        row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        if any(v is not None and str(v).strip() for v in row.values()):
            out.append(row)
    return out


def _audit_status(audit_rows: List[Dict[str, Any]]) -> str:
    statuses = [str(row.get("Status", "")).upper() for row in audit_rows]
    if any(s == "FAIL" for s in statuses):
        return "FAIL"
    if any(s == "REVIEW REQUIRED" for s in statuses):
        return "REVIEW REQUIRED"
    if any(s == "PASS" for s in statuses):
        return "PASS"
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# Slack notification
# ---------------------------------------------------------------------------
def _send_slack(webhook_url: str, record: Dict[str, Any]) -> Optional[str]:
    if not webhook_url or not webhook_url.strip():
        return None
    audit_status = _audit_status(record.get("audit_rows", []))
    status_emoji = {"completed": "✅", "failed": "🚨"}.get(record.get("status", ""), "❓")
    text = (
        f"{status_emoji} *GEP Billing Watcher* — run `{record['run_id']}`\n"
        f"*Status:* {record.get('status', '').upper()}\n"
        f"*Billing period:* {record.get('billing_period') or 'Unknown'}\n"
        f"*Audit:* {audit_status}\n"
        f"*Usage file:* {record.get('usage_file_name', 'N/A')}\n"
        f"*Outputs:* {record.get('outputs_dir', 'N/A')}"
    )
    payload = {"text": text}
    req = urllib.request.Request(
        webhook_url.strip(),
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        urllib.request.urlopen(req, timeout=15).read()
        return None
    except urllib.error.URLError as exc:
        return f"Slack notification failed: {exc}"


# ---------------------------------------------------------------------------
# Scheduled accrual execution
# ---------------------------------------------------------------------------
def _should_run_accrual(ledger: Dict[str, Any], accrual_day: int) -> Optional[date]:
    """Return the accrual month (as first-of-month date) if an accrual should run now, else None.

    Runs when today >= accrual_day and no successful accrual exists for this month yet.
    """
    today = datetime.now(timezone.utc).date()
    if today.day < accrual_day:
        return None
    accrual_month = date(today.year, today.month, 1)
    month_key = f"{accrual_month.year:04d}-{accrual_month.month:02d}"
    accruals = ledger.get("accruals", {})
    existing = accruals.get(month_key, {})
    if existing.get("status") == "completed":
        return None
    return accrual_month


def _execute_accrual_run(
    accrual_month: date,
    inputs_dir: Path,
    usage_dir: Path,
    accrual_output_dir: Path,
    slack_webhook: str,
    config_filename: str,
) -> Dict[str, Any]:
    """Run the accrual engine for the given month. Returns a record for the ledger."""
    month_key = f"{accrual_month.year:04d}-{accrual_month.month:02d}"
    started_at = datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ")
    record: Dict[str, Any] = {
        "month": month_key,
        "started_at_utc": started_at,
        "status": "running",
        "trigger": "watcher_schedule",
    }

    logger.info("=" * 60)
    logger.info("ACCRUAL RUN STARTING  [%s]", month_key)
    logger.info("Using prior-month usage to estimate %s revenue", month_key)
    logger.info("=" * 60)

    rules_path = (inputs_dir / config_filename).resolve()
    try:
        je_path, totals_path, support_path = run_accrual(
            accrual_month=accrual_month,
            usage_dir=usage_dir,
            rules_path=rules_path,
            output_dir=accrual_output_dir,
            logger=logger,
            save_billing_detail=True,
        )
        record.update({
            "status": "completed",
            "completed_at_utc": datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "je_path": str(je_path),
            "totals_path": str(totals_path),
            "support_workbook_path": str(support_path),
        })
        logger.info("ACCRUAL RUN COMPLETED  [%s]", month_key)
        logger.info("JE CSV:     %s", je_path)
        logger.info("Totals CSV: %s", totals_path)
    except Exception:
        tb = traceback.format_exc()
        record.update({
            "status": "failed",
            "completed_at_utc": datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "error": tb,
        })
        logger.error("ACCRUAL RUN FAILED  [%s]\n%s", month_key, tb)

    if slack_webhook:
        status_emoji = {"completed": "\u2705", "failed": "\U0001f6a8"}.get(record["status"], "\u2753")
        text = (
            f"{status_emoji} *GEP Billing Watcher* — accrual `{month_key}`\n"
            f"*Status:* {record['status'].upper()}\n"
            f"*Accrual month:* {month_key}\n"
            f"*JE file:* {record.get('je_path', 'N/A')}"
        )
        slack_err = _send_slack(slack_webhook, {"status": record["status"], "run_id": f"accrual-{month_key}", "billing_period": month_key, "usage_file_name": "prior-month", "outputs_dir": str(accrual_output_dir), "audit_rows": []})
        if slack_err:
            logger.warning(slack_err)

    return record


# ---------------------------------------------------------------------------
# Billing run execution
# ---------------------------------------------------------------------------
def _execute_billing_run(
    usage_file: Path,
    inputs_dir: Path,
    outputs_dir: Path,
    config_filename: str,
    slack_webhook: str,
) -> Dict[str, Any]:
    """Run the billing engine for a detected usage file and return a result record."""
    run_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{uuid4().hex[:8]}"
    created_at = datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ")
    file_hash = _file_sha256(usage_file)

    record: Dict[str, Any] = {
        "run_id": run_id,
        "created_at_utc": created_at,
        "status": "running",
        "usage_file_name": usage_file.name,
        "usage_file_path": str(usage_file),
        "usage_file_sha256": file_hash,
        "outputs_dir": str(outputs_dir),
        "trigger": "watcher",
    }

    logger.info("=" * 60)
    logger.info("BILLING RUN STARTING  [%s]", run_id)
    logger.info("Usage file: %s", usage_file.name)
    logger.info("=" * 60)

    try:
        run_billing_engine(
            inputs_dir=inputs_dir,
            outputs_dir=outputs_dir,
            usage_prefix=DEFAULT_USAGE_PREFIX,
            config_filename=config_filename,
            logger=logger,
        )

        history_dir = outputs_dir / "gep_billing_log"
        report_candidates = sorted(history_dir.glob("*_Master_Billing_Report.xlsx"))
        report_path = report_candidates[-1] if report_candidates else None
        manifest_candidates = sorted(history_dir.glob("*_run_manifest_*.json"))
        manifest_path = manifest_candidates[-1] if manifest_candidates else None
        netsuite_path = outputs_dir / "gep_netsuite_invoice_import.csv"
        partner_dir = outputs_dir / "gep_partner_details"

        billing_period = None
        if report_path is not None:
            stem = report_path.stem
            if "_" in stem:
                billing_period = stem.split("_")[0]

        audit_rows = _read_audit_controls(report_path) if report_path else []
        audit = _audit_status(audit_rows)

        record.update({
            "status": "completed",
            "completed_at_utc": datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "billing_period": billing_period,
            "master_billing_report": str(report_path) if report_path else "",
            "netsuite_import_file": str(netsuite_path) if netsuite_path.exists() else "",
            "partner_details_folder": str(partner_dir) if partner_dir.exists() else "",
            "run_manifest": str(manifest_path) if manifest_path else "",
            "audit_status": audit,
            "audit_rows": audit_rows,
        })

        logger.info("=" * 60)
        logger.info("BILLING RUN COMPLETED  [%s]", run_id)
        logger.info("Billing period: %s", billing_period)
        logger.info("Audit status: %s", audit)
        logger.info("Master report: %s", report_path)
        logger.info("NetSuite CSV: %s", netsuite_path)
        logger.info("=" * 60)

    except Exception:
        tb = traceback.format_exc()
        record.update({
            "status": "failed",
            "completed_at_utc": datetime.now(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "error": tb,
        })
        logger.error("BILLING RUN FAILED  [%s]\n%s", run_id, tb)

    slack_err = _send_slack(slack_webhook, record)
    if slack_err:
        logger.warning(slack_err)
        record["slack_error"] = slack_err

    return record


# ---------------------------------------------------------------------------
# Core polling loop
# ---------------------------------------------------------------------------
def run_watcher(
    inputs_dir: Path,
    outputs_dir: Path,
    config_filename: str,
    poll_interval: int,
    slack_webhook: str,
    dry_run: bool = False,
    accrual_day: int = DEFAULT_ACCRUAL_DAY,
    usage_dir: Optional[Path] = None,
    accrual_output_dir: Optional[Path] = None,
    disable_accrual: bool = False,
) -> None:
    """Poll inputs_dir for new usage files and run billing when found.

    Also auto-runs accruals on the ``accrual_day`` of each month using
    the prior month's usage file (same basis as a manual CLI accrual).
    """
    _ensure_state_dir()
    ledger = _load_ledger()

    if usage_dir is None:
        usage_dir = DEFAULT_V3_USAGE_DIR
    if accrual_output_dir is None:
        accrual_output_dir = DEFAULT_ACCRUAL_OUTPUT_DIR

    logger.info("=" * 60)
    logger.info("GEP BILLING WATCHER STARTED")
    logger.info("  Inputs:        %s", inputs_dir)
    logger.info("  Outputs:       %s", outputs_dir)
    logger.info("  Rules file:    %s", config_filename)
    logger.info("  Poll interval: %d seconds (%d minutes)", poll_interval, poll_interval // 60)
    logger.info("  Slack:         %s", "configured" if slack_webhook else "not configured")
    logger.info("  Dry run:       %s", dry_run)
    logger.info("  Accrual:       %s (day %d)", "disabled" if disable_accrual else "enabled", accrual_day)
    logger.info("  Usage dir:     %s", usage_dir)
    logger.info("  Ledger:        %s", PROCESSED_LEDGER)
    logger.info("  Processed so far: %d file(s)", len(ledger["processed"]))
    logger.info("=" * 60)

    while not _shutdown_requested:
        try:
            # --- Usage file processing (existing behavior) ---
            candidates = _find_usage_candidates(inputs_dir, DEFAULT_USAGE_PREFIX)

            for usage_file in candidates:
                file_key = usage_file.name

                if file_key in ledger["processed"]:
                    continue

                if not _is_file_stable(usage_file):
                    continue

                file_hash = _file_sha256(usage_file)

                hash_already_processed = any(
                    entry.get("sha256") == file_hash
                    for entry in ledger["processed"].values()
                )
                if hash_already_processed:
                    logger.warning(
                        "File %s has same hash as a previously processed file — skipping (possible rename).",
                        usage_file.name,
                    )
                    ledger["processed"][file_key] = {
                        "sha256": file_hash,
                        "skipped_duplicate_hash": True,
                        "detected_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    }
                    _save_ledger(ledger)
                    continue

                if dry_run:
                    logger.info("[DRY RUN] Would process: %s (sha256: %s)", usage_file.name, file_hash[:16])
                    ledger["processed"][file_key] = {
                        "sha256": file_hash,
                        "dry_run": True,
                        "detected_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    }
                    _save_ledger(ledger)
                    continue

                record = _execute_billing_run(
                    usage_file=usage_file,
                    inputs_dir=inputs_dir,
                    outputs_dir=outputs_dir,
                    config_filename=config_filename,
                    slack_webhook=slack_webhook,
                )

                ledger["processed"][file_key] = {
                    "sha256": file_hash,
                    "run_id": record.get("run_id"),
                    "status": record.get("status"),
                    "billing_period": record.get("billing_period"),
                    "processed_at_utc": record.get("completed_at_utc") or record.get("created_at_utc"),
                }
                _save_ledger(ledger)

            # --- Scheduled accrual ---
            if not disable_accrual and not dry_run:
                accrual_month = _should_run_accrual(ledger, accrual_day)
                if accrual_month is not None:
                    month_key = f"{accrual_month.year:04d}-{accrual_month.month:02d}"
                    logger.info("Accrual trigger: today >= day %d and no completed accrual for %s", accrual_day, month_key)
                    acc_record = _execute_accrual_run(
                        accrual_month=accrual_month,
                        inputs_dir=inputs_dir,
                        usage_dir=usage_dir,
                        accrual_output_dir=accrual_output_dir,
                        slack_webhook=slack_webhook,
                        config_filename=config_filename,
                    )
                    if "accruals" not in ledger:
                        ledger["accruals"] = {}
                    ledger["accruals"][month_key] = acc_record
                    _save_ledger(ledger)

        except Exception:
            logger.exception("Error during poll cycle — will retry next cycle.")

        if _shutdown_requested:
            break

        logger.debug("Sleeping %d seconds until next poll...", poll_interval)
        wake_at = time.monotonic() + poll_interval
        while time.monotonic() < wake_at:
            if _shutdown_requested:
                break
            time.sleep(min(5, wake_at - time.monotonic()))

    logger.info("Watcher stopped.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Poll for new GEP usage files and run billing automatically."
    )
    parser.add_argument(
        "--inputs-dir",
        default=str(DEFAULT_INPUTS_DIR),
        help="Directory to watch for gepusage*.csv files (default: Google Drive inputs).",
    )
    parser.add_argument(
        "--outputs-dir",
        default=str(DEFAULT_OUTPUTS_DIR),
        help="Output directory for billing results (default: Google Drive outputs).",
    )
    parser.add_argument(
        "--config-file",
        default=DEFAULT_CONFIG_FILE,
        help="Rules workbook filename inside inputs directory.",
    )
    parser.add_argument(
        "--poll-interval",
        type=int,
        default=DEFAULT_POLL_INTERVAL,
        help="Seconds between folder checks (default: 1800 = 30 minutes).",
    )
    parser.add_argument(
        "--slack-webhook",
        default=os.getenv("BILLING_SLACK_WEBHOOK", ""),
        help="Slack webhook URL for notifications (or set BILLING_SLACK_WEBHOOK env var).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Detect files but don't actually run billing.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
    )
    parser.add_argument(
        "--reset-ledger",
        action="store_true",
        help="Clear the processed-files ledger and start fresh.",
    )
    parser.add_argument(
        "--accrual-day",
        type=int,
        default=DEFAULT_ACCRUAL_DAY,
        help=f"Day of month to auto-run accruals (default: {DEFAULT_ACCRUAL_DAY}).",
    )
    parser.add_argument(
        "--usage-dir",
        default=str(DEFAULT_V3_USAGE_DIR),
        help="Directory with prior-month usage CSVs for accrual (default: v3 query exports on Google Drive).",
    )
    parser.add_argument(
        "--accrual-output-dir",
        default=str(DEFAULT_ACCRUAL_OUTPUT_DIR),
        help="Output directory for accrual JE and totals CSVs.",
    )
    parser.add_argument(
        "--disable-accrual",
        action="store_true",
        help="Disable automatic accrual scheduling.",
    )
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s [watcher] %(message)s",
    )

    signal.signal(signal.SIGINT, _handle_signal)
    signal.signal(signal.SIGTERM, _handle_signal)

    if args.reset_ledger:
        _ensure_state_dir()
        if PROCESSED_LEDGER.exists():
            PROCESSED_LEDGER.unlink()
            logger.info("Ledger cleared.")

    inputs_dir = Path(args.inputs_dir).resolve()
    outputs_dir = Path(args.outputs_dir).resolve()
    usage_dir = Path(args.usage_dir).resolve()
    accrual_output_dir = Path(args.accrual_output_dir).resolve()

    if not inputs_dir.exists():
        logger.error("Inputs directory does not exist: %s", inputs_dir)
        return 2

    try:
        run_watcher(
            inputs_dir=inputs_dir,
            outputs_dir=outputs_dir,
            config_filename=args.config_file,
            poll_interval=args.poll_interval,
            slack_webhook=args.slack_webhook,
            dry_run=args.dry_run,
            accrual_day=args.accrual_day,
            usage_dir=usage_dir,
            accrual_output_dir=accrual_output_dir,
            disable_accrual=args.disable_accrual,
        )
    except BillingEngineError as exc:
        logger.error("%s", exc)
        return 2
    except Exception:
        logger.exception("Unexpected watcher failure.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
