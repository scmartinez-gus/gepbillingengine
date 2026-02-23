#!/usr/bin/env python3
"""Streamlit portal for running GEP billing and viewing results.

MVP features:
- Upload usage CSV
- Validate required structure before run
- Execute billing_engine on a per-run isolated workspace
- Show audit/control outcomes
- Download generated artifacts
- Persist run history
- Optional Slack notification webhook on completion/failure
"""

from __future__ import annotations

import io
import json
import logging
import os
import shutil
import traceback
import urllib.error
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from billing_engine import (
    DEFAULT_CONFIG_FILE,
    DEFAULT_INPUTS_DIR,
    DEFAULT_USAGE_PREFIX,
    parse_date,
    run_billing_engine,
)


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
LOGGER = logging.getLogger("billing_portal")

APP_ROOT = Path("outputs") / "portal_runs"
RUNS_META_DIR = APP_ROOT / "runs"
RUNS_ARTIFACTS_DIR = APP_ROOT / "artifacts"

REQUIRED_RULE_SHEETS = ["Pricing", "Minimums", "Config", "Mapping"]
REQUIRED_USAGE_COLS = ["for_month", "partner_id", "partner_name", "total_individual_users", "current_ach_speed"]


def ensure_dirs() -> None:
    """Create persistent directories used by portal."""
    RUNS_META_DIR.mkdir(parents=True, exist_ok=True)
    RUNS_ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)


def now_utc() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def new_run_id() -> str:
    return f"{now_utc().strftime('%Y%m%dT%H%M%SZ')}-{uuid4().hex[:8]}"


def run_record_path(run_id: str) -> Path:
    return RUNS_META_DIR / f"{run_id}.json"


def write_run_record(record: Dict[str, Any]) -> None:
    path = run_record_path(record["run_id"])
    with path.open("w", encoding="utf-8") as handle:
        json.dump(record, handle, indent=2, sort_keys=True)


def load_run_records() -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for path in sorted(RUNS_META_DIR.glob("*.json"), reverse=True):
        try:
            with path.open("r", encoding="utf-8") as handle:
                records.append(json.load(handle))
        except Exception:
            continue
    return records


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for col in df.columns:
        c = str(col).strip().lower()
        c = "".join(ch if ch.isalnum() else "_" for ch in c)
        c = "_".join(part for part in c.split("_") if part)
        renamed[col] = c
    return df.rename(columns=renamed)


def validate_usage_dataframe(df: pd.DataFrame) -> Tuple[List[str], List[str], Optional[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    normalized = normalize_cols(df)
    missing = [col for col in REQUIRED_USAGE_COLS if col not in normalized.columns]
    if missing:
        errors.append(f"Missing required usage columns: {missing}")

    billing_period = None
    if "for_month" in normalized.columns:
        for value in normalized["for_month"].tolist():
            d = parse_date(value)
            if d is not None:
                billing_period = f"{d.year:04d}.{d.month:02d}"
                break
        if billing_period is None:
            warnings.append("Could not determine billing month from FOR_MONTH.")

    if len(df) == 0:
        errors.append("Usage file has no rows.")

    return errors, warnings, billing_period


def validate_rules_workbook(rules_path: Path) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    if not rules_path.exists():
        errors.append(f"Rules workbook not found: {rules_path}")
        return errors, warnings
    try:
        xls = pd.ExcelFile(rules_path)
    except Exception as exc:
        errors.append(f"Failed to read rules workbook: {exc}")
        return errors, warnings
    missing_sheets = [sheet for sheet in REQUIRED_RULE_SHEETS if sheet not in xls.sheet_names]
    if missing_sheets:
        errors.append(f"Rules workbook missing required tabs: {missing_sheets}")
    return errors, warnings


def read_audit_controls(report_path: Path) -> List[Dict[str, Any]]:
    if not report_path.exists():
        return []
    try:
        wb = load_workbook(report_path, data_only=True)
    except Exception:
        return []
    if "Audit & Controls" not in wb.sheetnames:
        return []
    ws = wb["Audit & Controls"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        if values is None:
            continue
        row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        if any(v is not None and str(v).strip() for v in row.values()):
            out.append(row)
    return out


def zip_partner_details(partner_folder: Path, zip_path: Path) -> Optional[Path]:
    if not partner_folder.exists():
        return None
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in partner_folder.rglob("*"):
            if file_path.is_file():
                archive.write(file_path, arcname=file_path.relative_to(partner_folder))
    return zip_path


def status_from_audit_rows(audit_rows: List[Dict[str, Any]]) -> str:
    statuses = [str(row.get("Status", "")).upper() for row in audit_rows]
    if any(status == "FAIL" for status in statuses):
        return "FAIL"
    if any(status == "REVIEW REQUIRED" for status in statuses):
        return "REVIEW REQUIRED"
    if any(status == "PASS" for status in statuses):
        return "PASS"
    return "UNKNOWN"


def send_slack_notification(webhook_url: str, record: Dict[str, Any]) -> Optional[str]:
    if not webhook_url.strip():
        return None
    audit_status = status_from_audit_rows(record.get("audit_rows", []))
    text = (
        f"GEP Billing Run `{record['run_id']}` {record.get('status', '').upper()}\n"
        f"Billing period: {record.get('billing_period') or 'Unknown'}\n"
        f"Audit status: {audit_status}\n"
        f"Usage file: {record.get('usage_file_name', 'N/A')}"
    )
    payload = {"text": text}
    req = urllib.request.Request(
        webhook_url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        urllib.request.urlopen(req, timeout=10).read()
        return None
    except urllib.error.URLError as exc:
        return f"Slack notification failed: {exc}"


def execute_run(
    usage_name: str,
    usage_bytes: bytes,
    rules_path: Path,
    slack_webhook: str,
) -> Dict[str, Any]:
    run_id = new_run_id()
    created_at = now_utc().strftime("%Y-%m-%dT%H:%M:%SZ")

    run_root = RUNS_ARTIFACTS_DIR / run_id
    run_inputs = run_root / "inputs"
    run_outputs = run_root / "outputs"
    run_inputs.mkdir(parents=True, exist_ok=True)
    run_outputs.mkdir(parents=True, exist_ok=True)

    safe_usage_name = usage_name if usage_name.lower().endswith(".csv") else f"{usage_name}.csv"
    if not safe_usage_name.lower().startswith(DEFAULT_USAGE_PREFIX.lower()):
        safe_usage_name = f"{DEFAULT_USAGE_PREFIX}_{safe_usage_name}"
    usage_path = run_inputs / safe_usage_name
    usage_path.write_bytes(usage_bytes)

    copied_rules_path = run_inputs / DEFAULT_CONFIG_FILE
    shutil.copy2(rules_path, copied_rules_path)

    record: Dict[str, Any] = {
        "run_id": run_id,
        "created_at_utc": created_at,
        "status": "running",
        "usage_file_name": usage_name,
        "rules_file_path": str(rules_path),
        "run_root": str(run_root),
        "outputs_dir": str(run_outputs),
    }
    write_run_record(record)

    try:
        run_billing_engine(
            inputs_dir=run_inputs,
            outputs_dir=run_outputs,
            usage_prefix=DEFAULT_USAGE_PREFIX,
            config_filename=DEFAULT_CONFIG_FILE,
            logger=LOGGER,
        )
        history_dir = run_outputs / "gep_billing_log"
        report_candidates = sorted(history_dir.glob("*_Master_Billing_Report.xlsx"))
        report_path = report_candidates[-1] if report_candidates else None
        manifest_candidates = sorted(history_dir.glob("*_run_manifest_*.json"))
        manifest_path = manifest_candidates[-1] if manifest_candidates else None
        netsuite_path = run_outputs / "gep_netsuite_invoice_import.csv"
        partner_dir = run_outputs / "gep_partner_details"
        partner_zip_path = run_root / "partner_details.zip"
        zipped = zip_partner_details(partner_dir, partner_zip_path)

        billing_period = None
        if report_path is not None:
            stem = report_path.stem
            if "_" in stem:
                billing_period = stem.split("_")[0]

        audit_rows = read_audit_controls(report_path) if report_path else []
        record.update(
            {
                "status": "completed",
                "completed_at_utc": now_utc().strftime("%Y-%m-%dT%H:%M:%SZ"),
                "billing_period": billing_period,
                "master_billing_report": str(report_path) if report_path else "",
                "netsuite_import_file": str(netsuite_path) if netsuite_path.exists() else "",
                "partner_details_folder": str(partner_dir) if partner_dir.exists() else "",
                "partner_details_zip": str(zipped) if zipped else "",
                "run_manifest": str(manifest_path) if manifest_path else "",
                "audit_rows": audit_rows,
            }
        )
    except Exception:
        record.update(
            {
                "status": "failed",
                "completed_at_utc": now_utc().strftime("%Y-%m-%dT%H:%M:%SZ"),
                "error": traceback.format_exc(),
            }
        )

    slack_error = send_slack_notification(slack_webhook, record)
    if slack_error:
        record["slack_error"] = slack_error

    write_run_record(record)
    return record


def render_download_button(label: str, path_str: str, button_key: str) -> None:
    if not path_str:
        return
    path = Path(path_str)
    if not path.exists() or not path.is_file():
        return
    st.download_button(
        label,
        data=path.read_bytes(),
        file_name=path.name,
        mime="application/octet-stream",
        key=button_key,
    )


def render_run_details(record: Dict[str, Any], context_key: str) -> None:
    st.subheader(f"Run {record.get('run_id')}")
    st.write(
        {
            "status": record.get("status"),
            "created_at_utc": record.get("created_at_utc"),
            "completed_at_utc": record.get("completed_at_utc"),
            "billing_period": record.get("billing_period"),
            "usage_file_name": record.get("usage_file_name"),
        }
    )

    if record.get("status") == "failed":
        st.error("Run failed.")
        st.code(record.get("error", "No traceback available"))

    audit_rows = record.get("audit_rows", [])
    if audit_rows:
        st.markdown("**Audit & Controls**")
        st.dataframe(pd.DataFrame(audit_rows), use_container_width=True, hide_index=True)

    col1, col2, col3, col4 = st.columns(4)
    run_id = record.get("run_id", "unknown")
    key_prefix = f"{context_key}_{run_id}"
    with col1:
        render_download_button(
            "Master Billing Report",
            record.get("master_billing_report", ""),
            button_key=f"{key_prefix}_dl_master",
        )
    with col2:
        render_download_button(
            "NetSuite Import CSV",
            record.get("netsuite_import_file", ""),
            button_key=f"{key_prefix}_dl_netsuite",
        )
    with col3:
        render_download_button(
            "Partner Details ZIP",
            record.get("partner_details_zip", ""),
            button_key=f"{key_prefix}_dl_partner_zip",
        )
    with col4:
        render_download_button(
            "Run Manifest JSON",
            record.get("run_manifest", ""),
            button_key=f"{key_prefix}_dl_manifest",
        )

    if record.get("slack_error"):
        st.warning(record["slack_error"])


def main() -> None:
    st.set_page_config(page_title="GEP Billing Portal", page_icon=":bar_chart:", layout="wide")
    ensure_dirs()

    st.title("GEP Billing Import Portal")
    st.caption("Upload usage data, run billing, and retrieve outputs without manual folder digging.")

    default_rules = Path(DEFAULT_INPUTS_DIR) / DEFAULT_CONFIG_FILE
    rules_path_text = st.sidebar.text_input("Rules workbook path", value=str(default_rules))
    slack_webhook = st.sidebar.text_input(
        "Slack webhook URL (optional)",
        value=os.getenv("BILLING_SLACK_WEBHOOK", ""),
        type="password",
    )

    tab_run, tab_history = st.tabs(["Run Billing", "Run History"])

    with tab_run:
        st.markdown("### 1) Upload Usage CSV")
        usage_file = st.file_uploader("Usage file", type=["csv"], accept_multiple_files=False)

        rules_path = Path(rules_path_text)
        rules_errors, rules_warnings = validate_rules_workbook(rules_path)

        if rules_errors:
            for msg in rules_errors:
                st.error(msg)
        for msg in rules_warnings:
            st.warning(msg)

        usage_errors: List[str] = []
        usage_warnings: List[str] = []
        billing_period = None
        usage_bytes = b""

        if usage_file is not None:
            usage_bytes = usage_file.getvalue()
            try:
                preview_df = pd.read_csv(io.BytesIO(usage_bytes), dtype=object)
                usage_errors, usage_warnings, billing_period = validate_usage_dataframe(preview_df)
                st.markdown("### 2) Pre-run Validation")
                if billing_period:
                    st.success(f"Detected billing period: {billing_period}")
                for msg in usage_errors:
                    st.error(msg)
                for msg in usage_warnings:
                    st.warning(msg)
                with st.expander("Usage preview (first 25 rows)", expanded=False):
                    st.dataframe(preview_df.head(25), use_container_width=True, hide_index=True)
            except Exception as exc:
                usage_errors = [f"Could not parse usage CSV: {exc}"]
                st.error(usage_errors[0])

        run_disabled = bool(rules_errors) or bool(usage_errors) or usage_file is None

        st.markdown("### 3) Run")
        if st.button("Run Billing", type="primary", disabled=run_disabled, use_container_width=True):
            with st.spinner("Running billing engine..."):
                record = execute_run(
                    usage_name=usage_file.name,
                    usage_bytes=usage_bytes,
                    rules_path=rules_path,
                    slack_webhook=slack_webhook,
                )
            st.session_state["last_run_id"] = record["run_id"]
            if record["status"] == "completed":
                st.success("Run completed.")
            else:
                st.error("Run failed.")
            render_run_details(record, context_key="run_result")

        last_run_id = st.session_state.get("last_run_id")
        if last_run_id:
            st.markdown("---")
            st.caption("Most recent run in this session")
            last_path = run_record_path(last_run_id)
            if last_path.exists():
                with last_path.open("r", encoding="utf-8") as handle:
                    render_run_details(json.load(handle), context_key="last_run")

    with tab_history:
        st.markdown("### Run History")
        records = load_run_records()
        if not records:
            st.info("No runs yet.")
        else:
            table_rows = [
                {
                    "run_id": rec.get("run_id"),
                    "status": rec.get("status"),
                    "billing_period": rec.get("billing_period"),
                    "created_at_utc": rec.get("created_at_utc"),
                    "usage_file_name": rec.get("usage_file_name"),
                }
                for rec in records
            ]
            st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

            selected_id = st.selectbox("Select a run", options=[r["run_id"] for r in records])
            selected = next((r for r in records if r["run_id"] == selected_id), None)
            if selected:
                render_run_details(selected, context_key="history_selected")


if __name__ == "__main__":
    main()
